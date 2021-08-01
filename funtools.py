import numpy as np
import random
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font, NamedStyle, NumberFormatDescriptor, Color
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.formatting import rule
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule, FormulaRule, FormatObject
from openpyxl.workbook import protection
import string

desired_width = 320
np.set_printoptions(linewidth=desired_width)

# -------------------------------- Module Variables Global Variables ----------------------------- #

team_num = 32
rnds = 4
rinks = 0
teams = []
games = np.empty(1, dtype=np.int64)
grid = np.empty(1, dtype=np.int64)
my_count = 0
init_draw = np.empty(1, dtype=np.int64)
init_update = np.empty(1, dtype=np.int64)
header = np.empty(1, dtype='<U15')
final_draw = np.empty(1, dtype=np.int64)
random_draw = 0
a = np.empty(1,dtype=np.int64)



# ------------------------------------ Excel Styles ---------------------------------------------- #



# -------------------------------- Module Functions ---------------------------------- #


def my_fixture(no_of_teams):

    if len(no_of_teams) % 2 != 0:
        no_of_teams.append(0)  # if team number is odd - use 'day off' as fake team

    rotation = list(no_of_teams)       # copy the list
    fixtures = []
    full_draw = []

    for i in range(0, len(no_of_teams)-1):
        fixtures.append(rotation)
        rotation = [rotation[0]] + [rotation[-1]] + rotation[1:-1]

    for f in fixtures:
        n = len(f)
        full_draw.append(list(zip(f[0:int(n / 2)], reversed(f[int(n / 2):n]))))

    return full_draw

# --------------------------------------------------------------------------------------------------- #


def reshape():
    global games
    teams_r = range(0, team_num)
    for t in teams_r:
        teams.append(teams_r[t]+1)

    if random_draw == 1:
        random.shuffle(teams)

    games = np.reshape(my_fixture(teams), (team_num-1, team_num))
    return


# ---------------------------------------------------------------------------------------------- #

def possible(y, x, n, m):

    # print(y,x,n,m)

    for k in range(0, (rinks * 2)):  # 0 to 16
        if grid[y][k] == n or grid[y][k] == m:
            return False  # check in the row for n or m
    for i in range(0, rnds):  # 0 to 5
        if grid[i][x] == n or grid[i][x] == m:
            return False  # check in the column for m or n (has either team played on that rink)
        if grid[i][x+1] == n or grid[i][x+1] == m:
            return False  # check in the column for m or n (has either team played on that rink)
    return True
# y is the round number from the number of rounds set
# x = rink number from the number of rinks available (starts at 0 and counts by 2 - so 0 to 15 for 8 rinks


# ----------------------------------------------------------------------------------------------- #


def solve(level=0):

    global my_count
    my_count = my_count + 1

    if my_count > 1000000:
        print("There appears to be no solution, try a smaller number of rounds")
        # print(np.array(grid))
        return
    # print(my_count)
    for y in range(0, rnds):  # cycle through the number of rounds
        # print("Thinking!!!!")
        # print(games)
        z = 0
        for x in range(0, (rinks*2), 2):  # cycle through the number of rinks
            # LOOK AT ADDING AND X offset for rink seperation
            x = a[z]
            z = z + 1
            if grid[y][x] == 0 and grid[y][x+1] == 0:
                # print("This is y",y,"This is x",x)
                for n in range(0, len(games[y]), 2):    # cycle through every game in the draw in that round
                    # print("This is x", x)
                    if possible(y, x, games[y][n], games[y][n+1]):
                        grid[y][x] = games[y][n]
                        grid[y][x+1] = games[y][n+1]
                        # print(level)  # print levels of recursion
                        solve(level=level+1)
                        grid[y][x] = 0
                        grid[y][x + 1] = 0
                        #  print(level)

                return

    # print(np.array(grid))
    cv = np.array(grid)
    my_file = open('my_new_draw', 'wb')
    np.save(my_file, cv)
    my_file.close()
    # print(np.array(grid), file=filename)
    print(cv)
    # input("More?")
    arrange_export()

    exit()


# ---------------------------------------------------------------------------------------------- #

def create_draw():  # Used to create init_draw

    z = np.empty(team_num, dtype=np.int64)
    for t in range(1, team_num+1):
        # z.append(t)
        z[t-1] = t
    # print(z)
    t_arr = np.array(z)
    new_t = t_arr.reshape((team_num, 1))

    draw_grid = np.zeros((team_num, (rnds * 2)), dtype=np.int64)

    draw_grid = np.append(new_t, draw_grid, 1)

    return draw_grid
# create_draw()

# ----------------------------LOAD DRAW AND CREATE TOURNAMENT DRAW ------------------------------------------ #


def create_tournament_draw():
    my_file = open('my_new_draw', 'rb')
    t_draw = np.load(my_file)
    global init_draw

    for r in range(0, rnds):  # loop through each round

        for q1 in range(0, team_num):  # loop through each team in init_draw
            for p1 in range(0, team_num, 2):  # loop through the teams in t_draw starting at 0 in steps of 2
                if init_draw[q1][0] == t_draw[r][p1]:  # compare teams numbers to left team number of game
                    rink = divmod(p1, 2)
                    init_draw[q1][1 + 2*r] = rink[0]+1  # calculate and insert rink for game into init_draw
                    init_draw[q1][2 + 2*r] = t_draw[r][p1+1]
                    # when there is a match take the Vs team and insert into init_draw

        for q2 in range(0, team_num):  # loop through each team in init_draw
            for p2 in range(1, team_num, 2):  # loop through the teams in t_draw starting at 1 in steps of 2
                if init_draw[q2][0] == t_draw[r][p2]:  # compare teams numbers to right team number of game
                    rink = divmod(p2, 2)
                    init_draw[q2][1 + 2*r] = rink[0]+1  # calculate and insert rink for game into init_draw
                    init_draw[q2][2 + 2*r] = t_draw[r][p2-1]
                    # when there is a match take the Vs team and insert into init_draw
    return init_draw

# ---------------------------------- ADD For and Against columns to create_tournament_draw output ------------ #


def insert_for_against():
    global init_update

    test = np.zeros((team_num,  2), dtype=np.int64)

    for x in range(3, (rnds*4), 4):
        init_update = np.insert(init_update, [x], test, axis=1)
    # print(test)
    return


# ------------------------------------CREATE HEADER ---------------------------------------------------------- #


def make_header():
    global header
    header = ["Team No"]
    rnd_count = 1
    for x in range(1, rnds*4, 4):

        header.append("Rink")
        header.append("Rnd " + str(rnd_count) + " Vs")
        header.append('For')
        header.append('Against')
        rnd_count += 1

    header.append('Points')
    header.append('Tot For')
    header.append('Tot Against')
    header.append('Percentage')
    header.append('Standings')
    return header


# print(make_header())


# ----------------------------ADD HEADER TO DRAW AND EXPORT---------------------------------------------------- #


def add_header_export_draw():
    global final_draw
    final_draw = init_update
    # final_draw = np.append(header, init_update, axis=0)  # DOESNT add header due to mix of dtypes
    my_name = 'No of Teams ' + str(team_num) + ' - ' + 'No of Rnds ' + str(rnds)
    df = pd.DataFrame(final_draw)
    df.to_excel('output_data.xlsx', sheet_name=my_name, startcol=0, startrow=1, index=False, header=False)
    return final_draw

# ---------------------------------------- Calculate column positions for "FOR" ------------------------------ #


def excel_stuff():
    wb = load_workbook('output_data.xlsx')
    # wb._named_styles['Normal'].number_format = '#,##0.00'
    sheet = wb['No of Teams ' + str(team_num) + ' - ' + 'No of Rnds ' + str(rnds)]

    min_column = wb.active.min_column
    max_column = wb.active.max_column
    min_row = wb.active.min_row
    max_row = wb.active.max_row
    alphabet = list(string.ascii_uppercase)
    alpha_loop = (divmod(max_column, 26))[0]
    mod_alpha = []
    my_ary_for = []
    my_ary_against = []
    my_ary_header = []
    my_ary_points = []
    for_array = np.empty(rnds, dtype='<U10')
    against_array = np.empty(rnds, dtype='<U10')
    #   points_array = []
    # ------------------------------------ Excel Styles Defined -------------------------------------- #

    my_header_sty = NamedStyle(name="my_header_sty")
    my_header_sty.font=Font(bold=True, size=12, italic= True)
    bd = Side(style='medium', color='000000')
    my_header_sty.border=Border(left=bd, top=bd, right=bd, bottom=bd)
    my_header_sty.alignment  = Alignment(horizontal='center', vertical='center')
    my_header_sty.fill = PatternFill("solid", fgColor='00FFFFCC')

    my_for_sty = NamedStyle(name="my_for_sty")
    my_for_sty.font=Font(bold=False, size=11)
    bd = Side(style='thin', color='000000')
    bd1 = Side(style='medium', color='000000')
    my_for_sty.border=Border(left=bd1, top=bd, right=bd, bottom=bd)
    my_for_sty.alignment  = Alignment(horizontal='center', vertical='center')
    my_for_sty.fill = PatternFill("solid", fgColor='00CCFFCC')

    my_agnst_sty = NamedStyle(name="my_agnst_sty")
    my_agnst_sty.font=Font(bold=False, size=11)
    bd = Side(style='thin', color='000000')
    bd1 = Side(style='medium', color='000000')
    my_agnst_sty.border=Border(left=bd, top=bd, right=bd1, bottom=bd)
    my_agnst_sty.alignment  = Alignment(horizontal='center', vertical='center')
    my_agnst_sty.fill = PatternFill("solid", fgColor='00FFCCDD')

    my_results_sty = NamedStyle(name="my_results_sty")
    my_results_sty.font=Font(bold=False, size=12, italic= True)
    bd = Side(style='thin', color='000000')
    my_results_sty.border=Border(left=bd, top=bd, right=bd, bottom=bd)
    my_results_sty.alignment  = Alignment(horizontal='center', vertical='center')
    my_results_sty.fill = PatternFill("solid", fgColor='00FFFFCC')

    my_Vs_sty = NamedStyle(name="my_Vs_sty")
    my_Vs_sty.font=Font(bold=True, size=12, italic= False, color='00FF0000')
    bd = Side(style='thin', color='000000')
    my_Vs_sty.border=Border(left=bd, top=bd, right=bd, bottom=bd)
    my_Vs_sty.alignment  = Alignment(horizontal='center', vertical='center')
    # my_Vs_sty.fill = PatternFill("solid", fgColor='00C0C0C0')

    my_rink_sty = NamedStyle(name="my_rink_sty")
    my_rink_sty.font=Font(bold=True, size=12, italic= False, color='00000080')
    bd = Side(style='thin', color='000000')
    my_rink_sty.border=Border(left=bd, top=bd, right=bd, bottom=bd)
    my_rink_sty.alignment  = Alignment(horizontal='center', vertical='center')
    # my_rink_sty.fill = PatternFill("solid", fgColor='00C0C0C0')

    my_stand_sty = NamedStyle(name="my_stand_sty") # used for points as well
    my_stand_sty.font=Font(bold=True, size=13, italic= True, color='00000000')
    bd = Side(style='thin', color='000000')
    bd1 = Side(style='medium', color='000000')
    my_stand_sty.border=Border(left=bd, top=bd, right=bd1, bottom=bd)
    my_stand_sty.alignment  = Alignment(horizontal='center', vertical='center')
    # my_rink_sty.fill = PatternFill("solid", fgColor='00C0C0C0')

    my_points_sty = NamedStyle(name="my_points_sty") # used for points as well
    my_points_sty.font=Font(bold=True, size=13, italic= True, color='00000000')
    bd = Side(style='thin', color='000000')
    my_points_sty.border=Border(left=bd, top=bd, right=bd, bottom=bd)
    my_points_sty.alignment  = Alignment(horizontal='center', vertical='center')
    # my_rink_sty.fill = PatternFill("solid", fgColor='00C0C0C0')

    my_bottom_sty = NamedStyle(name="my_bottom_sty") # used for points as well
    bd = Side(style='medium', color='000000')
    my_bottom_sty.border=Border(top=bd)


    # ------------------------------ Excel workbook styles added ------------------------------- #
    # wb.add_named_style(my_header_sty)
    # wb.add_named_style(my_for_sty)
    # wb.add_named_style(my_agnst_sty)
    # wb.add_named_style(my_results_sty)
    # wb.add_named_style(my_Vs_sty)

    # ---------------------------- Protect the sheet -------------------------- #

    wb['No of Teams ' + str(team_num) + ' - ' + 'No of Rnds ' + str(rnds)].protection.sheet = True



    # ------------------------------------------------------------------------------------------ #

    for x in range(0, alpha_loop + 1):
        x_loop = alphabet[x]
        for a in range(0, len(alphabet)):
            mod_alpha.append(x_loop + alphabet[a])

    full_alpha = np.append(alphabet, mod_alpha)

    for r in range(3, (rnds*4), 4):
        my_ary_for.append(full_alpha[r])

    for p in range(4, (rnds*4)+1, 4):
        my_ary_against.append(full_alpha[p])

    for a in range(0, (rnds * 4)+6):  # excel alhpa for column names for header
        my_ary_header.append((full_alpha[a]))

    for d in range(3, (rnds*4), 4):   # for calculating the points
        my_ary_points.append(full_alpha[d])

    for w in range(2, team_num + 2): # sum the for scores
        for z in range(0, len(my_ary_for)):
            if z != len(my_ary_for)-1:
                for_array[z] = my_ary_for[z] + str(w) + ' + '
            else:
                for_array[z] = my_ary_for[z] + str(w)

        final_array = ' '.join(map(str, for_array))

        sheet[f'{full_alpha[(rnds*4)+2]}{w}'] = f'= SUM({final_array})'
        sheet[f'{full_alpha[(rnds*4)+2]}{w}'].alignment  = Alignment(horizontal='center', vertical='center')

    for w in range(2, team_num + 2): # sum the against scores
        for z in range(0, len(my_ary_against)):
            if z != len(my_ary_against)-1:
                against_array[z] = my_ary_against[z] + str(w) + ' + '
            else:
                against_array[z] = my_ary_against[z] + str(w)

        final_array_a = ' '.join(map(str, against_array))  # convert list to string

        sheet[f'{full_alpha[(rnds*4)+3]}{w}'] = f'= SUM({final_array_a})'
        sheet[f'{full_alpha[(rnds*4)+3]}{w}'].alignment  = Alignment(horizontal='center', vertical='center')

# this is the area to make the points formula ---------------------------------------------------#
    for w in range(2, team_num + 2):
        points_array = []
        for z in range(3, (rnds*4) + 1, 4):
            if z != (rnds*4)-1:
                for_e = full_alpha[z] + str(w)
                agst_e = full_alpha[z+1] + str(w)
                points_array.append('IF(' + for_e + '>' + agst_e + ',10,IF(AND(' + for_e + '='+agst_e + ','
                                    + for_e + '<>0),5,0))+')
            else:
                for_e = full_alpha[z] + str(w)
                agst_e = full_alpha[z+1] + str(w)
                points_array.append('IF(' + for_e + '>' + agst_e + ',10,IF(AND(' + for_e + '='+agst_e + ','
                                    + for_e + '<>0),5,0))')

        final_array_p = ' '.join(map(str, points_array))  # convert list to a single string
        #  print(final_array_p)
        sheet[f'{full_alpha[(rnds*4)+1]}{w}'] = f'={final_array_p}'  # SUM game results
        sheet[f'{full_alpha[(rnds*4)+4]}{w}'] = f'=IFERROR({full_alpha[(rnds*4)+2]}{w}/{full_alpha[(rnds*4)+3]}{w}' \
                                                f',0)'  # Percentage
        sheet[f'{full_alpha[(rnds*4)+5]}{w}'] = f'=RANK({full_alpha[(rnds*4)+1]}{w},${full_alpha[(rnds*4)+1]}$2:${full_alpha[(rnds*4)+1]}${team_num+1})' \
                                                f'+ COUNTIFS(${full_alpha[(rnds*4)+1]}$2:${full_alpha[(rnds*4)+1]}${team_num+1}' \
                                                f',${full_alpha[(rnds*4)+1]}{w},${full_alpha[(rnds*4)+4]}$2:${full_alpha[(rnds*4)+4]}${team_num+1}' \
                                                f',">"&${full_alpha[(rnds*4)+4]}{w})' # this calcualtes the standings
        sheet[f'{full_alpha[(rnds*4)+1]}{w}'].alignment  = Alignment(horizontal='center', vertical='center')
        sheet[f'{full_alpha[(rnds*4)+4]}{w}'].alignment  = Alignment(horizontal='center', vertical='center')

    for t in range(0, (rnds * 4)+6):
        # print(len(my_ary_header), len(header))

        head_slice =header[t]
        cell_pos = my_ary_header[t]+ '1'
        sheet[cell_pos]= head_slice
        sheet[cell_pos].style = my_header_sty
        sheet.column_dimensions [my_ary_header[t]].width = len(head_slice) + 3


    for x in range (1, max_column): # format draw info in excel (not teams, header or results)
        # print(full_alpha[x])
        for y in range (2, team_num+ 2):
            #print(y)
            sheet[full_alpha[x]+str(y)].alignment  = Alignment(horizontal='center', vertical='center')


    for x in range (0, 1): # format team list info in excel
        # print(full_alpha[x])
        for y in range (2, team_num+ 2):
            #print(y)
            sheet[full_alpha[x]+str(y)].style = my_header_sty

    for x in range (2, (rnds * 4) + 1,4): # format each Vs info in excel
        # print(full_alpha[x])
        for y in range (2, team_num+ 2):
            #print(y)
            sheet[full_alpha[x]+str(y)].style = my_Vs_sty

    for x in range (1, (rnds * 4) + 1,4): # format each Rink info in excel
        # print(full_alpha[x])
        for y in range (2, team_num+ 2):
            #print(y)
            sheet[full_alpha[x]+str(y)].style = my_rink_sty

    for x in range (3, rnds * 4,4): # format each games For info in excel
        # print(full_alpha[x])
        for y in range (2, team_num+ 2):
            #print(y)
            sheet[full_alpha[x]+str(y)].style = my_for_sty
            sheet[full_alpha[x]+str(y)].protection = Protection(locked=False)

    for x in range (4, (rnds * 4) + 1,4): # format each games Against info in excel
        # print(full_alpha[x])
        for y in range (2, team_num+ 2):
            #print(y)
            sheet[full_alpha[x]+str(y)].style = my_agnst_sty
            sheet[full_alpha[x]+str(y)].protection = Protection(locked=False)

    for x in range ((rnds * 4) + 2, (rnds * 4) + 5,1): # format Tot For/Tot Against/Percentage info in excel
        # print(full_alpha[x])
        for y in range (2, team_num+ 2):
            #print(y)
            sheet[full_alpha[x]+str(y)].style = my_results_sty

    for x in range ((rnds * 4)+ 4, (rnds * 4)+ 5,1): # format Percentage column
        # print(full_alpha[x])
        for y in range (2, team_num+ 2):
            #print(y)
            sheet[full_alpha[x]+str(y)].number_format = '0.00%'

    for x in range ((rnds * 4)+ 5, (rnds * 4)+ 6,1): # format standings and Points column
        # print(full_alpha[x])
        for y in range (2, team_num+ 2):
            #print(y)
            sheet[full_alpha[x]+str(y)].style=my_stand_sty
            sheet[full_alpha[x-4]+str(y)].style=my_points_sty

    for x in range (1, (rnds * 4)+ 6,1): # format boareder on last row
        # print(full_alpha[x])
        sheet[full_alpha[x]+str(max_row+1)].style=my_bottom_sty




    # -------------------------------- Set the conditional formatting for the points column -------------- #
    sheet.conditional_formatting.add(f'{full_alpha[(rnds*4)+1]}{2}:{full_alpha[(rnds*4)+1]}{team_num+1}',
                                     ColorScaleRule(start_type='min',start_color='00FFFFFF', end_type='max',
                                                            end_color='00FFD500'))
    # ------------------------------- Set the conditional formatting for the standings column ------------ #
    sheet.conditional_formatting.add(f'{full_alpha[(rnds*4)+5]}{2}:{full_alpha[(rnds*4)+5]}{team_num+1}',
                                     ColorScaleRule(start_type='min',start_color='00FFD500', end_type='max',
                                                    end_color='00FFFFFF'))

    return wb.save('output_data.xlsx')


# ------------------------------------------------------------------------------------------------------- #


def arrange_export():
    global init_draw
    global init_update
    global header
    init_draw = (create_draw())
    init_update = create_tournament_draw()
    # print(init_update.dtype)
    insert_for_against()
    header = make_header()
    add_header_export_draw()
    excel_stuff()
    return
# -------------------------------------------------------------------------------------------------------------- #


def make_my_draw(num_of_rnds=4, num_of_teams=32, draw_random=0):
    global rnds
    global team_num
    global grid
    global rinks
    global init_draw
    global init_update
    global header
    global random_draw
    global a


    if draw_random == 0:
        random_draw = 0
    else:
        random_draw = 1

    rnds = num_of_rnds
    team_num = num_of_teams
    if team_num % 2 != 0:
        team_num = team_num + 1
    rinks = int(team_num/2)
    grid = np.zeros((rnds, (rinks * 2)), dtype=np.int64)
    a = np.arange(0,(rinks * 2), 2,dtype=np.int64)
    random.shuffle(a)
    # print(a)
    reshape()
    solve()
    return

# ------------------------------------------------------------------------------------------------ #
